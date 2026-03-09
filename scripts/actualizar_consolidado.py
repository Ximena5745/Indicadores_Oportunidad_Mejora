"""
Script para actualizar Resultados Consolidados.xlsx
Versión 2 - con manejo correcto de indicadores monetarios/series

Tipos de indicadores:
  1. Porcentaje: resultado/meta son % → usar API resultado/meta directamente
  2. Valor unitario (kWh, m3, estudiantes): usar API resultado/meta o suma de series
  3. Monetario ($): extraer de variables (planeado/ejecutado) o series.meta/resultado
     - Pesos: mostrar en pesos (sin conversión, usar la escala ya en Historico)

Columnas formuladas (G,H,I,L,M,R): se escriben como fórmulas Excel

Fuentes (data/raw/):
  - API/{year}.xlsx        : datos de la API por año (2022-2025)
  - Kawak/{year}.xlsx      : datos Kawak por año (2021, 2025)
  - Indicadores por CMI.xlsx
  - Resultados Consolidados.xlsx  (solo lectura → base)

Salida (data/output/):
  - Resultados Consolidados.xlsx  (archivo actualizado)
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import pandas as pd
import numpy as np
import ast
import warnings
import calendar
import shutil
import openpyxl
from pathlib import Path

warnings.filterwarnings('ignore')

# ── Rutas ──────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).resolve().parent.parent

DATA_INPUT  = BASE_DIR / "data" / "raw"          # solo lectura
DATA_OUTPUT = BASE_DIR / "data" / "output"        # escritura
DATA_OUTPUT.mkdir(parents=True, exist_ok=True)

INPUT_FILE  = DATA_INPUT  / "Resultados Consolidados.xlsx"
OUTPUT_FILE = DATA_OUTPUT / "Resultados Consolidados.xlsx"

# Migración automática: si el output no existe pero sí el input, copiar
if not OUTPUT_FILE.exists() and INPUT_FILE.exists():
    shutil.copy2(str(INPUT_FILE), str(OUTPUT_FILE))
    print(f"INFO: Resultados Consolidados.xlsx copiado a data/output/")

# ── Palabras clave para identificar Meta vs Ejecucion en variables ──
KW_EJEC = ['real', 'ejecutado', 'recaudado', 'ahorrado', 'consumo', 'generado',
           'actual', 'logrado', 'obtenido', 'reportado', 'hoy']
KW_META = ['planeado', 'presupuestado', 'propuesto', 'programado', 'objetivo',
           'esperado', 'previsto', 'estimado', 'acumulado plan']

IDS_MONETARIOS_UNIT = None  # se determina dinámicamente


# ─────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────

def make_llave(id_val, fecha):
    try:
        id_str = str(id_val)
        if id_str.endswith('.0'):
            id_str = id_str[:-2]
        d = pd.to_datetime(fecha)
        return f"{id_str}-{d.year}-{str(d.month).zfill(2)}-{str(d.day).zfill(2)}"
    except Exception:
        return None


def ultimo_dia_mes(year, month):
    return calendar.monthrange(year, month)[1]


def fechas_por_periodicidad(periodicidad, year=2025):
    mapa = {
        'Mensual':    [12, 11, 10, 9, 8, 7, 6, 5, 4, 3, 2, 1],
        'Trimestral': [12, 9, 6, 3],
        'Semestral':  [12, 6],
        'Anual':      [12],
        'Bimestral':  [12, 10, 8, 6, 4, 2],
    }
    meses = mapa.get(periodicidad, [12])
    return [pd.Timestamp(year, m, ultimo_dia_mes(year, m)) for m in meses]


def limpiar_clasificacion(val):
    if isinstance(val, str):
        return (val.replace('Estrat&eacute;gico', 'Estratégico')
                   .replace('&eacute;', 'é')
                   .replace('&amp;', '&'))
    return val


def parse_json_safe(val):
    if pd.isna(val) or val == '' or val is None:
        return None
    try:
        return ast.literal_eval(str(val))
    except Exception:
        return None


def limpiar_html(val):
    """Elimina entidades HTML básicas."""
    if not isinstance(val, str):
        return val
    return (val.replace('&oacute;', 'ó').replace('&eacute;', 'é')
               .replace('&aacute;', 'á').replace('&iacute;', 'í')
               .replace('&uacute;', 'ú').replace('&ntilde;', 'ñ')
               .replace('&Eacute;', 'É').replace('&amp;', '&'))


def _id_str(val):
    """Normaliza un ID a string sin '.0' final."""
    s = str(val)
    if s.endswith('.0'):
        s = s[:-2]
    return s


# ─────────────────────────────────────────────────────────────────────
# EXTRACCIÓN DE META Y EJECUCION DESDE VARIABLES/SERIES
# ─────────────────────────────────────────────────────────────────────

def extraer_meta_ejec_variables(vars_list):
    """
    Dado el listado de variables JSON de la API, extrae Meta y Ejecucion.
    Usa keywords para distinguir el rol de cada variable.
    """
    if not vars_list:
        return None, None

    meta_val = None
    ejec_val = None

    for v in vars_list:
        nombre = str(v.get('nombre', '')).lower()
        valor  = v.get('valor', None)
        if valor is None or (isinstance(valor, float) and np.isnan(valor)):
            continue
        if any(kw in nombre for kw in KW_META) and meta_val is None:
            meta_val = valor
        elif any(kw in nombre for kw in KW_EJEC) and ejec_val is None:
            ejec_val = valor

    if meta_val is None and len(vars_list) >= 2:
        meta_val = vars_list[1].get('valor')
    if ejec_val is None and len(vars_list) >= 1:
        ejec_val = vars_list[0].get('valor')

    return meta_val, ejec_val


def extraer_meta_ejec_series(series_list):
    """Suma meta y resultado de todas las subseries."""
    if not series_list:
        return None, None
    sum_meta = 0.0
    sum_res  = 0.0
    has_meta = False
    has_res  = False
    for s in series_list:
        m = s.get('meta')
        r = s.get('resultado')
        if m is not None and not (isinstance(m, float) and np.isnan(m)):
            sum_meta += float(m)
            has_meta = True
        if r is not None and not (isinstance(r, float) and np.isnan(r)):
            sum_res += float(r)
            has_res = True
    return (sum_meta if has_meta else None), (sum_res if has_res else None)


def determinar_meta_ejec(row_api, hist_meta_escala):
    """
    Determina Meta y Ejecucion correctas para un registro de la API.
    hist_meta_escala: escala típica de Meta en Historico para este indicador.
    """
    resultado   = row_api.get('resultado')
    meta_api    = row_api.get('meta')
    vars_list   = parse_json_safe(row_api.get('variables'))
    series_list = parse_json_safe(row_api.get('series'))

    es_grande = (hist_meta_escala is not None and hist_meta_escala > 1000)
    api_es_porcentaje = (meta_api is not None and
                         not (isinstance(meta_api, float) and np.isnan(meta_api)) and
                         abs(float(meta_api)) <= 200)

    if es_grande and api_es_porcentaje:
        if vars_list:
            meta_v, ejec_v = extraer_meta_ejec_variables(vars_list)
            if ejec_v is not None:
                return meta_v, ejec_v, 'variables'
        if series_list:
            sum_m, sum_r = extraer_meta_ejec_series(series_list)
            if sum_r is not None:
                return sum_m, sum_r, 'series_sum'
        return None, None, 'skip'

    if (resultado is None or (isinstance(resultado, float) and np.isnan(resultado))):
        if series_list:
            sum_m, sum_r = extraer_meta_ejec_series(series_list)
            if sum_r is not None:
                return sum_m, sum_r, 'series_sum_fallback'
        return None, None, 'sin_resultado'

    return meta_api, resultado, 'api_directo'


# ─────────────────────────────────────────────────────────────────────
# CARGA Y NORMALIZACIÓN DE FUENTES
# ─────────────────────────────────────────────────────────────────────

def cargar_api(years=(2022, 2023, 2024, 2025)):
    frames = []
    for y in years:
        path = DATA_INPUT / "API" / f"{y}.xlsx"
        if not path.exists():
            continue
        df = pd.read_excel(path)
        df['año_archivo'] = y
        frames.append(df)
    if not frames:
        print("  ADVERTENCIA: No se encontraron archivos API en data/raw/API/")
        return pd.DataFrame(columns=['Id', 'Indicador', 'Proceso', 'Periodicidad',
                                     'Sentido', 'resultado', 'meta', 'fecha',
                                     'LLAVE', 'variables', 'series', 'analisis'])
    df = pd.concat(frames, ignore_index=True)
    df = df.dropna(subset=['fecha'])
    df['fecha'] = pd.to_datetime(df['fecha'])
    df['clasificacion'] = df['clasificacion'].apply(limpiar_clasificacion)
    df = df.rename(columns={
        'ID': 'Id', 'nombre': 'Indicador', 'proceso': 'Proceso',
        'frecuencia': 'Periodicidad', 'sentido': 'Sentido',
    })
    df['LLAVE'] = df.apply(lambda r: make_llave(r['Id'], r['fecha']), axis=1)
    return df


def cargar_kawak_old(years=(2021,)):
    """Solo usamos 2021 (API cubre 2022+)."""
    frames = []
    for y in years:
        path = DATA_INPUT / "Kawak" / f"{y}.xlsx"
        if not path.exists():
            continue
        df = pd.read_excel(path)
        df['año_archivo'] = y
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    df = df.dropna(subset=['fecha'])
    df['fecha'] = pd.to_datetime(df['fecha'])
    df['clasificacion'] = df['clasificacion'].apply(limpiar_clasificacion)
    df = df.rename(columns={
        'ID': 'Id', 'nombre': 'Indicador', 'proceso': 'Proceso',
        'sentido': 'Sentido',
    })
    if 'frecuencia' in df.columns:
        df = df.rename(columns={'frecuencia': 'Periodicidad'})
    elif 'Periodicidad' not in df.columns:
        df['Periodicidad'] = 'Mensual'
    df['LLAVE'] = df.apply(lambda r: make_llave(r['Id'], r['fecha']), axis=1)
    return df


def cargar_kawak_2025():
    path = DATA_INPUT / "Kawak" / "2025.xlsx"
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_excel(path)
    rename_map = {}
    for col in df.columns:
        if 'Clasificaci' in col:
            rename_map[col] = 'clasificacion'
        elif 'Meta' in col and 'ltimo' in col:
            rename_map[col] = 'Meta'
        elif 'Tipo de variable' in col:
            rename_map[col] = 'Tipo_variable'
        elif 'Tipo de calculo' in col:
            rename_map[col] = 'TipoCalculo'
        elif 'Nombre variable' in col:
            rename_map[col] = 'NombreVar'
    df = df.rename(columns=rename_map)
    if 'clasificacion' not in df.columns:
        df['clasificacion'] = ''
    df['clasificacion'] = df['clasificacion'].apply(limpiar_clasificacion)

    periodo_cols = [c for c in df.columns if str(c).startswith('Periodo ')]

    df_global = (df[df.get('NombreVar', pd.Series()).str.contains('Consolidado Global', na=False)].copy()
                 if 'NombreVar' in df.columns else pd.DataFrame())
    ids_sin_global = set(df['Id']) - set(df_global['Id'])
    if ids_sin_global:
        extra = df[df['Id'].isin(ids_sin_global)].drop_duplicates('Id')
        df_global = pd.concat([df_global, extra], ignore_index=True)

    records = []
    for _, row in df_global.iterrows():
        periodicidad = row.get('Periodicidad', 'Mensual')
        fechas = fechas_por_periodicidad(periodicidad, 2025)
        tipo_calc = row.get('TipoCalculo', '')
        for i, col in enumerate(periodo_cols):
            if i >= len(fechas):
                break
            valor = row.get(col)
            if pd.isna(valor):
                continue
            records.append({
                'Id':           row['Id'],
                'Indicador':    limpiar_html(str(row.get('Indicador', ''))),
                'clasificacion':row['clasificacion'],
                'Proceso':      row.get('Proceso', ''),
                'Periodicidad': periodicidad,
                'Sentido':      row.get('Sentido', ''),
                'TipoCalculo':  tipo_calc,
                'Meta':         row.get('Meta', np.nan),
                'resultado':    valor,
                'meta':         row.get('Meta', np.nan),
                'fecha':        fechas[i],
                'fuente':       'Kawak2025',
            })
    if not records:
        return pd.DataFrame()
    df_k = pd.DataFrame(records)
    df_k['LLAVE'] = df_k.apply(lambda r: make_llave(r['Id'], r['fecha']), axis=1)
    return df_k


# ─────────────────────────────────────────────────────────────────────
# METADATOS MAESTROS DESDE KAWAK Y CMI
# ─────────────────────────────────────────────────────────────────────

def cargar_metadatos_kawak():
    """
    Consolida metadatos de todos los archivos Kawak.
    Prioridad: 2025 > 2024 > 2023 > 2022 > 2021.
    """
    meta = {}
    for y in [2021, 2022, 2023, 2024]:
        path = DATA_INPUT / "Kawak" / f"{y}.xlsx"
        if not path.exists():
            continue
        df = pd.read_excel(path)
        id_col  = 'ID' if 'ID' in df.columns else 'Id'
        per_col = ('frecuencia' if 'frecuencia' in df.columns else
                   'Periodicidad' if 'Periodicidad' in df.columns else None)
        for _, row in df.drop_duplicates(id_col).iterrows():
            id_val = row.get(id_col)
            if pd.isna(id_val):
                continue
            ids = _id_str(id_val)
            meta[ids] = {
                'nombre':        limpiar_html(str(row.get('nombre', row.get('Indicador', '')))),
                'clasificacion': limpiar_clasificacion(str(row.get('clasificacion',
                                                                    row.get('Clasificacion', '')))),
                'proceso':       limpiar_html(str(row.get('proceso', row.get('Proceso', '')))),
                'periodicidad':  str(row.get(per_col, '')) if per_col else '',
                'sentido':       str(row.get('sentido', row.get('Sentido', ''))),
                'tipo_calculo':  '',
            }

    path25 = DATA_INPUT / "Kawak" / "2025.xlsx"
    if path25.exists():
        df25 = pd.read_excel(path25)
        clas_col = next((c for c in df25.columns if 'Clasificaci' in c), None)
        tc_col   = next((c for c in df25.columns if 'Tipo de calculo' in c), None)
        for _, row in df25.drop_duplicates('Id').iterrows():
            id_val = row.get('Id')
            if pd.isna(id_val):
                continue
            ids = _id_str(id_val)
            meta[ids] = {
                'nombre':        limpiar_html(str(row.get('Indicador', ''))),
                'clasificacion': limpiar_clasificacion(
                                     str(row.get(clas_col, '')) if clas_col else ''),
                'proceso':       limpiar_html(str(row.get('Proceso', ''))),
                'periodicidad':  str(row.get('Periodicidad', '')),
                'sentido':       str(row.get('Sentido', '')),
                'tipo_calculo':  str(row.get(tc_col, '')) if tc_col else '',
            }
    return meta


def cargar_metadatos_cmi():
    """
    Carga metadatos desde 'Indicadores por CMI.xlsx' (hoja Worksheet).
    Usado como fallback cuando el ID no está en Kawak.
    """
    path = DATA_INPUT / "Indicadores por CMI.xlsx"
    if not path.exists():
        return {}
    try:
        df = pd.read_excel(path, sheet_name='Worksheet')
    except Exception:
        return {}

    clas_col = next((c for c in df.columns if 'Clasificaci' in c), None)
    meta = {}
    for _, row in df.iterrows():
        id_val = row.get('Id')
        if pd.isna(id_val):
            continue
        ids = _id_str(id_val)
        meta[ids] = {
            'nombre':        limpiar_html(str(row.get('Indicador', ''))),
            'clasificacion': limpiar_clasificacion(
                                 str(row.get(clas_col, '')) if clas_col else ''),
            'proceso':       limpiar_html(str(row.get('Subproceso', ''))),
            'periodicidad':  str(row.get('Periodicidad', '')),
            'sentido':       str(row.get('Sentido', '')),
            'tipo_calculo':  '',
        }
    return meta


# ─────────────────────────────────────────────────────────────────────
# CATÁLOGO DE INDICADORES
# ─────────────────────────────────────────────────────────────────────

def construir_catalogo(df_api, df_hist=None,
                       metadatos_kawak=None, metadatos_cmi=None):
    """
    Construye tabla maestra con metadatos de TODOS los indicadores.
    Preserva TipoCalculo y Asociacion editados por el usuario en el output anterior.
    """
    if metadatos_kawak is None:
        metadatos_kawak = {}
    if metadatos_cmi is None:
        metadatos_cmi = {}

    user_data = {}
    if OUTPUT_FILE.exists():
        try:
            xl = pd.ExcelFile(OUTPUT_FILE)
            if 'Catalogo Indicadores' in xl.sheet_names:
                df_ex = pd.read_excel(OUTPUT_FILE, sheet_name='Catalogo Indicadores')
                for _, row in df_ex.iterrows():
                    ids = _id_str(row['Id'])
                    user_data[ids] = {
                        'TipoCalculo': row.get('TipoCalculo', ''),
                        'Asociacion':  row.get('Asociacion', ''),
                    }
        except Exception:
            pass

    all_ids = {}

    if len(df_api) > 0:
        df_last_api = df_api.sort_values('fecha').groupby('Id').last().reset_index()
        for c in ['Indicador', 'clasificacion', 'Proceso', 'Periodicidad', 'Sentido', 'Tipo', 'estado']:
            if c not in df_last_api.columns:
                df_last_api[c] = ''
        for _, row in df_last_api.iterrows():
            ids = _id_str(row['Id'])
            all_ids[ids] = {
                'Id':          row['Id'],
                'Indicador':   limpiar_html(str(row['Indicador'])),
                'Clasificacion': limpiar_clasificacion(str(row['clasificacion'])),
                'Proceso':     limpiar_html(str(row['Proceso'])),
                'Periodicidad':str(row['Periodicidad']),
                'Sentido':     str(row['Sentido']),
                'Tipo_API':    str(row['Tipo']),
                'Estado':      str(row['estado']),
                'Fuente':      'API',
            }

    if df_hist is not None and len(df_hist) > 0:
        df_hc = df_hist.copy()
        df_hc['Fecha'] = pd.to_datetime(df_hc['Fecha'])
        df_hc_last = df_hc.sort_values('Fecha').groupby('Id').last().reset_index()
        col_ind  = next((c for c in ['Indicador', 'nombre'] if c in df_hc_last.columns), None)
        col_proc = 'Proceso' if 'Proceso' in df_hc_last.columns else None
        col_per  = 'Periodicidad' if 'Periodicidad' in df_hc_last.columns else None
        col_sent = 'Sentido' if 'Sentido' in df_hc_last.columns else None
        col_clas = next((c for c in ['Clasificacion', 'clasificacion'] if c in df_hc_last.columns), None)
        for _, row in df_hc_last.iterrows():
            ids = _id_str(row['Id'])
            if ids not in all_ids:
                all_ids[ids] = {
                    'Id':          row['Id'],
                    'Indicador':   limpiar_html(str(row[col_ind])) if col_ind else '',
                    'Clasificacion': limpiar_clasificacion(str(row[col_clas])) if col_clas else '',
                    'Proceso':     limpiar_html(str(row[col_proc])) if col_proc else '',
                    'Periodicidad':str(row[col_per]) if col_per else '',
                    'Sentido':     str(row[col_sent]) if col_sent else '',
                    'Tipo_API':    '',
                    'Estado':      'Historico',
                    'Fuente':      'Historico',
                }

    def _clean(v):
        return '' if (v is None or str(v).strip() in ('', 'nan', 'None')) else str(v).strip()

    rows = []
    for ids, base in all_ids.items():
        kw  = metadatos_kawak.get(ids, {})
        cmi = metadatos_cmi.get(ids, {})

        nombre        = _clean(kw.get('nombre'))        or _clean(cmi.get('nombre'))        or base['Indicador']
        clasificacion = _clean(kw.get('clasificacion')) or _clean(cmi.get('clasificacion')) or base['Clasificacion']
        proceso       = _clean(kw.get('proceso'))       or _clean(cmi.get('proceso'))       or base['Proceso']
        periodicidad  = _clean(kw.get('periodicidad'))  or _clean(cmi.get('periodicidad'))  or base['Periodicidad']
        sentido       = _clean(kw.get('sentido'))       or _clean(cmi.get('sentido'))       or base['Sentido']

        ud = user_data.get(ids, {})
        tipo_calculo = _clean(ud.get('TipoCalculo')) or _clean(kw.get('tipo_calculo', ''))
        asociacion   = _clean(ud.get('Asociacion', ''))

        rows.append({
            'Id':             base['Id'],
            'Indicador':      nombre,
            'Clasificacion':  clasificacion,
            'Proceso':        proceso,
            'Periodicidad':   periodicidad,
            'Sentido':        sentido,
            'Tipo_API':       base['Tipo_API'],
            'Estado':         base['Estado'],
            'Fuente':         base['Fuente'],
            'TipoCalculo':    tipo_calculo,
            'Asociacion':     asociacion,
            'Formato_Valores':'Porcentaje',
        })

    df_cat = pd.DataFrame(rows)

    def sort_key(id_val):
        try:
            return (0, float(str(id_val)))
        except Exception:
            return (1, str(id_val))

    df_cat = df_cat.sort_values('Id', key=lambda col: col.map(sort_key))
    return df_cat.reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────
# SERIES, VARIABLES, ANÁLISIS
# ─────────────────────────────────────────────────────────────────────

def expandir_series(df_api):
    rows = []
    for _, r in df_api.iterrows():
        parsed = parse_json_safe(r.get('series'))
        if not parsed:
            continue
        for s in parsed:
            row_base = {
                'Id':              r['Id'],
                'Indicador':       limpiar_html(str(r.get('Indicador', r.get('nombre', '')))),
                'Proceso':         r.get('Proceso', ''),
                'Periodicidad':    r.get('Periodicidad', ''),
                'Sentido':         r.get('Sentido', ''),
                'fecha':           r['fecha'],
                'LLAVE':           r['LLAVE'],
                'serie_nombre':    limpiar_html(str(s.get('nombre', ''))),
                'serie_meta':      s.get('meta'),
                'serie_resultado': s.get('resultado'),
            }
            for v in s.get('variables', []):
                row_base[f"var_{v.get('simbolo', 'X')}"] = v.get('valor')
            rows.append(row_base)
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def expandir_analisis(df_api):
    rows = []
    for _, r in df_api.iterrows():
        analisis = r.get('analisis', '')
        if pd.isna(analisis) or not str(analisis).strip():
            continue
        partes = str(analisis).split(' | ', 2)
        rows.append({
            'Id':             r['Id'],
            'Indicador':      limpiar_html(str(r.get('Indicador', ''))),
            'Proceso':        r.get('Proceso', ''),
            'fecha':          r['fecha'],
            'LLAVE':          r['LLAVE'],
            'analisis_fecha': partes[0].strip() if len(partes) > 0 else '',
            'analisis_autor': partes[1].strip() if len(partes) > 1 else '',
            'analisis_texto': limpiar_html(partes[2].strip() if len(partes) > 2 else str(analisis).strip()),
        })
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────
# SIGNOS DEL CONSOLIDADO EXISTENTE
# ─────────────────────────────────────────────────────────────────────

def obtener_signos(df_hist, df_sem, df_cierres):
    signos = {}
    for df, col_ms_candidates, col_es_candidates in [
        (df_hist,    ['Meta Signo', 'Meta s'], ['Ejecución Signo', 'Ejecucion Signo', 'Ejecución s']),
        (df_sem,     ['Meta Signo', 'Meta s'], ['Ejecucion Signo', 'Ejecución s']),
        (df_cierres, ['Meta Signo', 'Meta s'], ['Ejecucion Signo', 'Ejecución s']),
    ]:
        col_ms = next((c for c in col_ms_candidates if c in df.columns), None)
        col_es = next((c for c in col_es_candidates if c in df.columns), None)
        col_dm = next((c for c in ['Decimales_Meta', 'Decimales'] if c in df.columns), None)
        col_de = next((c for c in ['Decimales_Ejecucion', 'DecimalesEje'] if c in df.columns), None)

        for _, row in df.sort_values('Fecha').iterrows():
            id_str = str(row['Id'])
            signos[id_str] = {
                'meta_signo': row.get(col_ms, '%') if col_ms else '%',
                'ejec_signo': row.get(col_es, '%') if col_es else '%',
                'dec_meta':   row.get(col_dm, 0) if col_dm else 0,
                'dec_ejec':   row.get(col_de, 0) if col_de else 0,
            }
    return signos


# ─────────────────────────────────────────────────────────────────────
# FÓRMULAS EXCEL PARA FILAS NUEVAS
# ─────────────────────────────────────────────────────────────────────

def formula_G(r): return f"=YEAR(F{r})"
def formula_H(r): return f'=PROPER(TEXT(F{r},"mmmm"))'
def formula_I(r): return (
    f'=IF(OR(H{r}="Enero",H{r}="Febrero",H{r}="Marzo",'
    f'H{r}="Abril",H{r}="Mayo",H{r}="Junio"),'
    f'G{r}&"-1",'
    f'IF(OR(H{r}="Julio",H{r}="Agosto",H{r}="Septiembre",'
    f'H{r}="Octubre",H{r}="Noviembre",H{r}="Diciembre"),'
    f'G{r}&"-2"))'
)
def formula_L(r): return (
    f'=IFERROR(IF(E{r}="Positivo",'
    f'MIN(MAX(K{r}/J{r},0),1.3),'
    f'MIN(MAX(J{r}/K{r},0),1.3)),"")'
)
def formula_M(r): return (
    f'=IFERROR(IF(E{r}="Positivo",'
    f'MAX(K{r}/J{r},0),'
    f'MAX(J{r}/K{r},0)),"")'
)
def formula_R(r): return (
    f'=A{r}&"-"&YEAR(F{r})&"-"'
    f'&IF(LEN(MONTH(F{r}))=1,"0"&MONTH(F{r}),MONTH(F{r}))'
    f'&"-"&IF(LEN(DAY(F{r}))=1,"0"&DAY(F{r}),DAY(F{r}))'
)


def escribir_filas(ws, filas, signos, start_row=None):
    """Escribe filas con fórmulas en las columnas G,H,I,L,M,R."""
    r = start_row if start_row else ws.max_row + 1
    for fila in filas:
        id_str = str(fila.get('Id', ''))
        sg = signos.get(id_str, {'meta_signo': '%', 'ejec_signo': '%',
                                  'dec_meta': 0, 'dec_ejec': 0})

        fecha_val = fila.get('fecha')
        if isinstance(fecha_val, pd.Timestamp):
            fecha_val = fecha_val.to_pydatetime().date()

        meta = fila.get('Meta')
        ejec = fila.get('Ejecucion')

        def nan2none(v):
            return None if v is None or (isinstance(v, float) and np.isnan(v)) else v

        ws.cell(r, 1).value  = fila.get('Id')
        ws.cell(r, 2).value  = fila.get('Indicador', '')
        ws.cell(r, 3).value  = fila.get('Proceso', '')
        ws.cell(r, 4).value  = fila.get('Periodicidad', '')
        ws.cell(r, 5).value  = fila.get('Sentido', '')
        ws.cell(r, 6).value  = fecha_val
        ws.cell(r, 6).number_format = 'YYYY-MM-DD'
        ws.cell(r, 7).value  = formula_G(r)
        ws.cell(r, 8).value  = formula_H(r)
        ws.cell(r, 9).value  = formula_I(r)
        ws.cell(r, 10).value = nan2none(meta)
        ws.cell(r, 11).value = nan2none(ejec)
        ws.cell(r, 12).value = formula_L(r)
        ws.cell(r, 12).number_format = '0.00%'
        ws.cell(r, 13).value = formula_M(r)
        ws.cell(r, 13).number_format = '0.00%'
        ws.cell(r, 14).value = sg['meta_signo']
        ws.cell(r, 15).value = sg['ejec_signo']
        ws.cell(r, 16).value = sg['dec_meta']
        ws.cell(r, 17).value = sg['dec_ejec']
        ws.cell(r, 18).value = formula_R(r)
        r += 1
    return r - 1


def escribir_hoja_nueva(wb, nombre, df):
    if nombre in wb.sheetnames:
        del wb[nombre]
    ws = wb.create_sheet(nombre)
    for j, col in enumerate(df.columns, 1):
        ws.cell(1, j).value = col
    for i, (_, row) in enumerate(df.iterrows(), 2):
        for j, col in enumerate(df.columns, 1):
            val = row[col]
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime().date()
            elif isinstance(val, float) and np.isnan(val):
                val = None
            ws.cell(i, j).value = val


# ─────────────────────────────────────────────────────────────────────
# PREPARAR NUEVOS REGISTROS PARA CADA HOJA
# ─────────────────────────────────────────────────────────────────────

def construir_registros_para_hoja(df_fuente, llaves_existentes, hist_escalas,
                                   modo='historico', id_año_dic_existentes=None):
    """
    df_fuente: DataFrame normalizado (API o Kawak)
    hist_escalas: dict {id: meta_promedio_existente}
    modo: 'historico' | 'semestral' | 'cierres'
    id_año_dic_existentes: set de (str(Id), año) que YA tienen cierre de diciembre
    """
    registros = []
    skipped   = 0

    df = df_fuente.copy()

    if modo == 'semestral':
        df = df[df['fecha'].dt.month.isin([6, 12])]
        df = df[df['fecha'] == df['fecha'].apply(
            lambda d: pd.Timestamp(d.year, d.month,
                                   calendario_ultimo_dia(d.year, d.month)))]

    elif modo == 'cierres':
        df['año'] = df['fecha'].dt.year
        df['mes'] = df['fecha'].dt.month
        df['prioridad'] = df['mes'].apply(lambda m: 0 if m == 12 else 1)
        df = (df.sort_values(['Id', 'año', 'prioridad', 'fecha'],
                              ascending=[True, True, True, False])
                .groupby(['Id', 'año'])
                .first()
                .reset_index())
        if id_año_dic_existentes:
            df['key'] = df.apply(lambda r: (str(r['Id']), int(r['año'])), axis=1)
            df = df[~df['key'].isin(id_año_dic_existentes)]
            df = df.drop(columns=['key'])

    df = df[~df['LLAVE'].isin(llaves_existentes)]
    df = df.dropna(subset=['LLAVE'])

    for _, row in df.iterrows():
        id_val = row['Id']
        id_num = pd.to_numeric(id_val, errors='coerce')
        hist_meta_escala = hist_escalas.get(id_num) or hist_escalas.get(str(id_val))

        if 'fuente' in row and row.get('fuente') == 'Kawak2025':
            meta   = row.get('Meta')
            ejec   = row.get('resultado')
            fuente = 'Kawak2025'
        else:
            meta, ejec, fuente = determinar_meta_ejec(row.to_dict(), hist_meta_escala)

        if fuente == 'skip':
            skipped += 1
            continue

        registros.append({
            'Id':          id_val,
            'Indicador':   limpiar_html(str(row.get('Indicador', ''))),
            'Proceso':     row.get('Proceso', ''),
            'Periodicidad':row.get('Periodicidad', ''),
            'Sentido':     row.get('Sentido', ''),
            'fecha':       row['fecha'],
            'Meta':        meta,
            'Ejecucion':   ejec,
            'LLAVE':       row['LLAVE'],
        })

    return registros, skipped


def calendario_ultimo_dia(year, month):
    return calendar.monthrange(year, month)[1]


# ─────────────────────────────────────────────────────────────────────
# CORRECCIÓN DE CIERRES (1 por año, preferir diciembre)
# ─────────────────────────────────────────────────────────────────────

def corregir_cierres_wb(wb):
    """
    En Consolidado Cierres: para cada Id+Año con múltiples registros,
    dejar solo el de diciembre (o el último si no hay diciembre).
    """
    if 'Consolidado Cierres' not in wb.sheetnames:
        return
    ws = wb['Consolidado Cierres']

    data = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is None:
            continue
        data.append({
            'row_idx': row[0].row,
            'Id':       row[0].value,
            'Fecha':    row[5].value,
        })

    if not data:
        return

    df_c = pd.DataFrame([{'row_idx': d['row_idx'], 'Id': d['Id'],
                           'Fecha': pd.to_datetime(d['Fecha'], errors='coerce')} for d in data])
    df_c = df_c.dropna(subset=['Fecha'])
    df_c['año'] = df_c['Fecha'].dt.year
    df_c['mes'] = df_c['Fecha'].dt.month

    dup = df_c.groupby(['Id', 'año']).size().reset_index(name='cnt')
    dup_multi = dup[dup['cnt'] > 1]
    filas_a_borrar = set()

    for _, dr in dup_multi.iterrows():
        subset = df_c[(df_c['Id'] == dr['Id']) & (df_c['año'] == dr['año'])]
        tiene_dic = subset[subset['mes'] == 12]
        if len(tiene_dic) > 0:
            no_dic = subset[subset['mes'] != 12]
            filas_a_borrar.update(no_dic['row_idx'].tolist())
            if len(tiene_dic) > 1:
                keep = tiene_dic.sort_values('Fecha').iloc[-1]['row_idx']
                filas_a_borrar.update(
                    tiene_dic[tiene_dic['row_idx'] != keep]['row_idx'].tolist()
                )
        else:
            keep = subset.sort_values('Fecha').iloc[-1]['row_idx']
            filas_a_borrar.update(
                subset[subset['row_idx'] != keep]['row_idx'].tolist()
            )

    for r_idx in sorted(filas_a_borrar, reverse=True):
        ws.delete_rows(r_idx)

    print(f"  Cierres: {len(filas_a_borrar)} filas duplicadas eliminadas")


# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("ACTUALIZANDO RESULTADOS CONSOLIDADOS - v2")
    print("=" * 65)
    print(f"\n  Entradas : {DATA_INPUT}")
    print(f"  Salida   : {DATA_OUTPUT}")

    # ── 1. Cargar fuentes ──────────────────────────────────────────
    print("\n[1] Cargando fuentes de datos...")
    df_api = cargar_api()
    print(f"  API (2022-2025): {len(df_api):,} registros")

    df_kawak21 = cargar_kawak_old((2021,))
    print(f"  Kawak 2021: {len(df_kawak21):,} registros")

    df_kawak25 = cargar_kawak_2025()
    print(f"  Kawak 2025: {len(df_kawak25):,} registros")

    cols_base = ['Id', 'Indicador', 'Proceso', 'Periodicidad', 'Sentido',
                 'resultado', 'meta', 'fecha', 'LLAVE', 'variables', 'series', 'analisis']
    for c in cols_base:
        for df_ in [df_api, df_kawak21]:
            if c not in df_.columns:
                df_[c] = np.nan

    partes = [df_api[cols_base]] if len(df_api) > 0 else []
    if len(df_kawak21) > 0:
        partes.append(df_kawak21[cols_base])

    if partes:
        df_fuente_api = pd.concat(partes, ignore_index=True)
        df_fuente_api = df_fuente_api.drop_duplicates('LLAVE', keep='first')
        df_fuente_api = df_fuente_api.dropna(subset=['LLAVE'])
    else:
        df_fuente_api = pd.DataFrame(columns=cols_base)

    print(f"  Fuente unificada API+Kawak21: {len(df_fuente_api):,} registros")

    # ── 2. Verificar archivo base ──────────────────────────────────
    if not INPUT_FILE.exists() and not OUTPUT_FILE.exists():
        print(f"\n  ERROR: No se encontró {INPUT_FILE}")
        print(f"  Coloca 'Resultados Consolidados.xlsx' en data/raw/")
        return

    print("\n[2] Cargando Resultados Consolidados...")
    source = OUTPUT_FILE if OUTPUT_FILE.exists() else INPUT_FILE
    df_hist    = pd.read_excel(source, sheet_name='Consolidado Historico')
    df_sem     = pd.read_excel(source, sheet_name='Consolidado Semestral')
    df_cierres = pd.read_excel(source, sheet_name='Consolidado Cierres')

    for df_ in [df_hist, df_sem, df_cierres]:
        df_['Fecha'] = pd.to_datetime(df_['Fecha'], errors='coerce')

    df_hist['Meta_num'] = pd.to_numeric(df_hist['Meta'], errors='coerce')
    hist_escalas = df_hist.groupby('Id')['Meta_num'].median().to_dict()
    print(f"  Escalas cargadas para {len(hist_escalas)} indicadores")

    signos = obtener_signos(df_hist, df_sem, df_cierres)
    print(f"  Signos cargados para {len(signos)} indicadores")

    llave_hist    = set(df_hist['LLAVE'].dropna().astype(str))
    llave_sem     = set(df_sem['LLAVE'].dropna().astype(str))
    col_llave_c   = 'Llave' if 'Llave' in df_cierres.columns else 'LLAVE'
    llave_cierres = set(df_cierres[col_llave_c].dropna().astype(str))

    df_cierres['_mes'] = pd.to_datetime(df_cierres['Fecha'], errors='coerce').dt.month
    df_cierres['_año'] = pd.to_datetime(df_cierres['Fecha'], errors='coerce').dt.year
    id_año_dic = set(
        df_cierres[df_cierres['_mes'] == 12]
        .apply(lambda r: (str(r['Id']), int(r['_año'])), axis=1)
    )

    print(f"  Historico: {len(df_hist):,} | Semestral: {len(df_sem):,} | Cierres: {len(df_cierres):,}")
    print(f"  Cierres dic. existentes: {len(id_año_dic):,} combinaciones Id+Año")

    # ── 3. Metadatos maestros Kawak + CMI ─────────────────────────
    print("\n[3] Cargando metadatos maestros (Kawak > CMI)...")
    meta_kawak = cargar_metadatos_kawak()
    meta_cmi   = cargar_metadatos_cmi()
    print(f"  Kawak: {len(meta_kawak)} IDs | CMI: {len(meta_cmi)} IDs")

    def _apply_meta(row, field, fallback):
        ids = _id_str(row['Id'])
        v = (meta_kawak.get(ids, {}).get(field) or
             meta_cmi.get(ids, {}).get(field) or '').strip()
        return v if v and v not in ('nan', 'None') else fallback(row)

    if len(df_fuente_api) > 0:
        df_fuente_api['Indicador']    = df_fuente_api.apply(
            lambda r: _apply_meta(r, 'nombre',       lambda r: limpiar_html(str(r['Indicador']))), axis=1)
        df_fuente_api['Periodicidad'] = df_fuente_api.apply(
            lambda r: _apply_meta(r, 'periodicidad', lambda r: str(r['Periodicidad'])), axis=1)
        df_fuente_api['Proceso']      = df_fuente_api.apply(
            lambda r: _apply_meta(r, 'proceso',      lambda r: str(r['Proceso'])), axis=1)

    if len(df_kawak25) > 0:
        df_kawak25['Indicador']    = df_kawak25.apply(
            lambda r: _apply_meta(r, 'nombre',       lambda r: limpiar_html(str(r['Indicador']))), axis=1)
        df_kawak25['Periodicidad'] = df_kawak25.apply(
            lambda r: _apply_meta(r, 'periodicidad', lambda r: str(r['Periodicidad'])), axis=1)

    # ── 4. Series / Analisis ──────────────────────────────────────
    print("\n[4] Expandiendo series y análisis...")
    df_series   = expandir_series(df_api) if len(df_api) > 0 else pd.DataFrame()
    df_analisis = expandir_analisis(df_api) if len(df_api) > 0 else pd.DataFrame()
    print(f"  Series: {len(df_series):,} | Análisis: {len(df_analisis):,}")

    # ── 5. Catálogo ───────────────────────────────────────────────
    print("\n[5] Construyendo catálogo de indicadores...")
    df_cat = construir_catalogo(df_api if len(df_api) > 0 else pd.DataFrame(),
                                df_hist,
                                metadatos_kawak=meta_kawak,
                                metadatos_cmi=meta_cmi)
    print(f"  Catálogo: {len(df_cat):,} indicadores")

    # ── 6. Abrir workbook ─────────────────────────────────────────
    print("\n[6] Copiando base a outputs y abriendo workbook...")
    shutil.copy(str(INPUT_FILE if INPUT_FILE.exists() else OUTPUT_FILE), str(OUTPUT_FILE))
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    # ── 7. Historico ──────────────────────────────────────────────
    print("\n[7] Calculando registros nuevos para Historico...")
    regs_hist = []
    skip_hist = 0
    if len(df_fuente_api) > 0:
        regs_hist, skip_hist = construir_registros_para_hoja(
            df_fuente_api, llave_hist, hist_escalas, modo='historico')
    print(f"  Registros nuevos: {len(regs_hist):,} | Omitidos: {skip_hist}")

    if len(df_kawak25) > 0:
        llaves_ya = llave_hist | {r['LLAVE'] for r in regs_hist}
        regs_k25, sk25 = construir_registros_para_hoja(
            df_kawak25, llaves_ya, hist_escalas, modo='historico')
        regs_hist += regs_k25
        print(f"  + Kawak 2025: {len(regs_k25):,} adicionales (omitidos: {sk25})")

    regs_hist.sort(key=lambda x: (str(x['Id']), x['fecha']))
    if regs_hist and 'Consolidado Historico' in wb.sheetnames:
        ultima = escribir_filas(wb['Consolidado Historico'], regs_hist, signos)
        print(f"  Historico: ultima fila = {ultima}")

    # ── 8. Semestral ──────────────────────────────────────────────
    print("\n[8] Calculando registros nuevos para Semestral...")
    regs_sem = []
    if len(df_fuente_api) > 0:
        regs_sem, skip_sem = construir_registros_para_hoja(
            df_fuente_api, llave_sem, hist_escalas, modo='semestral')
        print(f"  Semestral nuevos: {len(regs_sem):,} | Omitidos: {skip_sem}")

    regs_sem.sort(key=lambda x: (str(x['Id']), x['fecha']))
    if regs_sem and 'Consolidado Semestral' in wb.sheetnames:
        ultima = escribir_filas(wb['Consolidado Semestral'], regs_sem, signos)
        print(f"  Semestral: ultima fila = {ultima}")

    # ── 9. Cierres ────────────────────────────────────────────────
    print("\n[9] Calculando registros nuevos para Cierres...")
    regs_cierres = []
    if len(df_fuente_api) > 0:
        regs_cierres, skip_c = construir_registros_para_hoja(
            df_fuente_api, llave_cierres, hist_escalas,
            modo='cierres', id_año_dic_existentes=id_año_dic)
        print(f"  Cierres nuevos: {len(regs_cierres):,} | Omitidos: {skip_c}")

    regs_cierres.sort(key=lambda x: (str(x['Id']), x['fecha']))
    if regs_cierres and 'Consolidado Cierres' in wb.sheetnames:
        ultima = escribir_filas(wb['Consolidado Cierres'], regs_cierres, signos)
        print(f"  Cierres: ultima fila = {ultima}")

    # ── 9b. Corrección duplicados Cierres ─────────────────────────
    print("\n[9b] Corrigiendo duplicados en Consolidado Cierres...")
    corregir_cierres_wb(wb)

    # ── 10. Hojas nuevas ──────────────────────────────────────────
    print("\n[10] Escribiendo hojas auxiliares...")
    if len(df_series) > 0:
        escribir_hoja_nueva(wb, 'Desglose Series', df_series)
        print(f"  Desglose Series: {len(df_series):,} filas")
    if len(df_analisis) > 0:
        escribir_hoja_nueva(wb, 'Desglose Analisis', df_analisis)
        print(f"  Desglose Analisis: {len(df_analisis):,} filas")

    escribir_hoja_nueva(wb, 'Catalogo Indicadores', df_cat)
    print(f"  Catalogo Indicadores: {len(df_cat):,} filas")

    if len(df_fuente_api) > 0:
        df_base = df_fuente_api[['Id', 'Indicador', 'Proceso', 'Periodicidad',
                                  'Sentido', 'fecha', 'resultado', 'meta', 'LLAVE']].copy()
        df_base['fecha'] = df_base['fecha'].dt.date
        escribir_hoja_nueva(wb, 'Base Normalizada', df_base)
        print(f"  Base Normalizada: {len(df_base):,} filas")

    # ── 11. Guardar ───────────────────────────────────────────────
    print(f"\nGuardando: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("[OK] Guardado exitosamente.")

    print("\n" + "=" * 65)
    print("RESUMEN:")
    print(f"  Historico:  +{len(regs_hist):,} nuevas filas")
    print(f"  Semestral:  +{len(regs_sem):,} nuevas filas")
    print(f"  Cierres:    +{len(regs_cierres):,} nuevas filas")
    print(f"  Hojas aux.: Desglose Series, Desglose Analisis,")
    print(f"              Catalogo Indicadores, Base Normalizada")
    print("=" * 65)


if __name__ == '__main__':
    main()
