#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
"""
generar_reporte.py
==================
Procesa 'lmi_reporte.xlsx' (descargado de plataforma externa) y genera
'Seguimiento_Reporte.xlsx' con:

  1. Hoja "Seguimiento"  -> copia completa + columna "Revisar"
  2. Hojas por periodicidad (Mensual, Bimestral, Trimestral, Semestral, Anual)
     - Columnas Periodo X renombradas a fechas reales (dd/mm/aaaa)
     - Columna "Reportado"          (período más reciente con/sin dato)
     - Columna "Estado del indicador" (últimos 2 períodos)

CONFIGURACIÓN
─────────────
  RUTA_ORIGEN      : ruta al archivo fuente .xlsx
  RUTA_SALIDA      : ruta de salida .xlsx
  COLUMNA_REVISAR  : columna usada para detectar inicio de nuevo indicador
  FECHA_REFERENCIA : fecha de corte; el Periodo 1 se asigna al período más
                     reciente ≤ esta fecha según cada periodicidad.
"""

import os
from datetime import date, timedelta
from typing import Optional  # Compatible con Python < 3.10

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuración ──────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

RUTA_ORIGEN = os.path.join(BASE_DIR, "data", "raw", "lmi_reporte.xlsx")
RUTA_SALIDA = os.path.join(BASE_DIR, "data", "output", "Seguimiento_Reporte.xlsx")

# Archivo de respaldo Kawak (puede no existir; se omite sin error si falta).
RUTA_KAWAK  = os.path.join(BASE_DIR, "data", "raw", "Fuentes Consolidadas", "Consolidado_API_Kawak.xlsx")

# Nombres de columnas en indicadores_kawak.xlsx
KAWAK_COL_ID     = "Id"
KAWAK_COL_FECHA  = "Fecha"
KAWAK_COL_RESULT = "Resultado"

# Columna cuyo cambio entre filas consecutivas marca un nuevo indicador único.
COLUMNA_REVISAR = "Id"

# Fecha de corte del reporte.
FECHA_REFERENCIA = date(2025, 12, 31)

# ── Paleta de colores ──────────────────────────────────────────────────────────
C_HEADER    = "1F4E79"
C_SI        = "C6EFCE"
C_NO        = "FFCCCC"
C_PENDIENTE = "FFEB9C"
C_REVISAR1  = "DDEEFF"

# ── Fechas de fin de período ───────────────────────────────────────────────────

def _ultimo_dia(year: int, month: int) -> date:
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - timedelta(days=1)


def _retroceder(year: int, mes_actual: int, meses_ciclo: list) -> tuple:
    idx = meses_ciclo.index(mes_actual)
    if idx == 0:
        return year - 1, meses_ciclo[-1]
    return year, meses_ciclo[idx - 1]


def get_period_dates(periodicidad: str, n: int = 13) -> list:
    ref = FECHA_REFERENCIA
    dates = []

    if periodicidad == "Mensual":
        y, m = ref.year, ref.month
        for _ in range(n):
            dates.append(_ultimo_dia(y, m))
            m -= 1
            if m == 0:
                m, y = 12, y - 1

    elif periodicidad == "Bimestral":
        ciclo = [2, 4, 6, 8, 10, 12]
        y, m = ref.year, ref.month
        previos = [x for x in ciclo if x <= m]
        cur = max(previos) if previos else 12
        if not previos:
            y -= 1
        for _ in range(n):
            dates.append(_ultimo_dia(y, cur))
            y, cur = _retroceder(y, cur, ciclo)

    elif periodicidad == "Trimestral":
        ciclo = [3, 6, 9, 12]
        y, m = ref.year, ref.month
        previos = [x for x in ciclo if x <= m]
        cur = max(previos) if previos else 12
        if not previos:
            y -= 1
        for _ in range(n):
            dates.append(_ultimo_dia(y, cur))
            y, cur = _retroceder(y, cur, ciclo)

    elif periodicidad == "Semestral":
        ciclo = [6, 12]
        y, m = ref.year, ref.month
        previos = [x for x in ciclo if x <= m]
        cur = max(previos) if previos else 12
        if not previos:
            y -= 1
        for _ in range(n):
            dates.append(_ultimo_dia(y, cur))
            y, cur = _retroceder(y, cur, ciclo)

    elif periodicidad == "Anual":
        y = ref.year
        for i in range(n):
            dates.append(date(y - i, 12, 31))

    else:
        dates = [None] * n

    return dates


# ── Leer archivo fuente ────────────────────────────────────────────────────────
# CAMBIO: se reemplazó leer_xls (basada en xlrd, solo para .xls legacy)
# por leer_xlsx usando pandas + openpyxl, compatible con archivos .xlsx actuales.

def leer_xlsx(ruta: str) -> pd.DataFrame:
    """
    Lee lmi_reporte.xlsx usando pandas con motor openpyxl.
    - keep_default_na=False evita que textos como 'N/A' se conviertan en NaN.
    - na_values=[""] trata celdas vacías como NaN.
    - dtype=str carga todo como texto para evitar conversiones automáticas
      no deseadas (p.ej. Ids numéricos leídos como float).
    """
    df = pd.read_excel(
        ruta,
        engine="openpyxl",
        keep_default_na=False,
        na_values=[""],
        dtype=str,          # Leer todo como texto; se limpian tipos después
    )
    # Limpiar nombres de columna con espacios extra
    df.columns = [str(c).strip() for c in df.columns]
    # Reemplazar 'nan' (string) producido por dtype=str en celdas vacías
    df = df.where(df != "nan", other=None)
    df = df.where(df != "NaN", other=None)
    return df


# ── Lógica de columnas "Revisar", "Reportado", "Estado" ───────────────────────

def agregar_revisar(df: pd.DataFrame, col: str) -> pd.DataFrame:
    df = df.copy()
    valores = df[col].tolist()
    revisar = [1] + [
        0 if valores[i] == valores[i - 1] else 1
        for i in range(1, len(valores))
    ]
    df["Revisar"] = revisar
    return df


def _tiene_dato(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return s not in ("", "-", "nan", "NaN", "None")


def _id_normalizar(x) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except (TypeError, ValueError):
        pass
    try:
        f = float(x)
        return str(int(f)) if f == int(f) else str(f)
    except (ValueError, TypeError):
        return str(x).strip()


def _tiene_dato_kawak(v) -> bool:
    try:
        if pd.isna(v):
            return False
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    return s not in ("", "-", "nan", "NaN", "None")


def _detectar_col(df: pd.DataFrame, candidatos: list) -> Optional[str]:
    # CAMBIO: tipo de retorno usa Optional[str] en vez de str | None
    # para compatibilidad con Python 3.9 y anteriores.
    col_lower = {c.lower(): c for c in df.columns}
    for cand in candidatos:
        found = col_lower.get(cand.lower())
        if found:
            return found
    return None


def leer_kawak(ruta: str) -> dict:
    if not os.path.exists(ruta):
        print(f"    INFO: {ruta} no encontrado — se omite cruce Kawak.")
        return {}

    try:
        df = pd.read_excel(ruta, engine="openpyxl",
                           keep_default_na=False, na_values=[""])
    except Exception as exc:
        print(f"    ADVERTENCIA kawak: no se pudo leer el archivo: {exc}")
        return {}

    df.columns = [str(c).strip() for c in df.columns]

    col_id  = _detectar_col(df, [KAWAK_COL_ID,     "Id", "ID", "Codigo", "Código", "codigo"])
    col_fec = _detectar_col(df, [KAWAK_COL_FECHA,  "Fecha", "Periodo", "Period", "FechaPeriodo", "Mes"])
    col_res = _detectar_col(df, [KAWAK_COL_RESULT, "Resultado", "Valor", "Value", "Result", "Dato"])

    if not col_id or not col_fec or not col_res:
        print(f"    ADVERTENCIA kawak: columnas requeridas no encontradas.\n"
              f"    Disponibles: {df.columns.tolist()}")
        return {}

    lookup = {}
    omitidas = 0
    for _, row in df.iterrows():
        kid = _id_normalizar(row[col_id])
        if not kid:
            continue

        fecha_raw = row[col_fec]
        resultado = row[col_res]

        try:
            if isinstance(fecha_raw, (pd.Timestamp, date)):
                ts = pd.Timestamp(fecha_raw)
            elif isinstance(fecha_raw, (int, float)) and not pd.isna(fecha_raw):
                ts = pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(fecha_raw))
            else:
                ts = pd.Timestamp(str(fecha_raw).strip())
            if pd.isna(resultado):
                continue
            lookup[(kid, ts.year, ts.month)] = resultado
        except Exception:
            omitidas += 1

    if omitidas:
        print(f"    INFO kawak: {omitidas} filas con fecha no parseable fueron omitidas.")

    print(f"    Kawak cargado: {len(lookup)} registros (Id × período).")
    return lookup


def enriquecer_desde_kawak(df_p: pd.DataFrame, kawak: dict,
                            col_p1: str, col_p2: Optional[str],  # CAMBIO: Optional[str]
                            fecha_p1, fecha_p2) -> tuple:
    if not kawak:
        return df_p, 0

    df_p = df_p.copy()
    n_act = 0

    for i, row in df_p.iterrows():
        if row.get("Estado del indicador") != "Pendiente de reporte":
            continue

        kid      = _id_normalizar(row.get("Id", ""))
        cambiado = False

        if fecha_p1 and not _tiene_dato(row.get(col_p1, "")):
            res1 = kawak.get((kid, fecha_p1.year, fecha_p1.month))
            if res1 is not None and _tiene_dato_kawak(res1):
                df_p.at[i, col_p1] = res1
                cambiado = True

        if col_p2 and fecha_p2 and not _tiene_dato(row.get(col_p2, "")):
            res2 = kawak.get((kid, fecha_p2.year, fecha_p2.month))
            if res2 is not None and _tiene_dato_kawak(res2):
                df_p.at[i, col_p2] = res2
                cambiado = True

        if cambiado:
            p1_ok = _tiene_dato(df_p.at[i, col_p1])
            df_p.at[i, "Reportado"]            = "Sí" if p1_ok else "No"
            df_p.at[i, "Estado del indicador"] = (
                "Reportado" if p1_ok else "Pendiente de reporte"
            )
            if p1_ok:
                n_act += 1

    return df_p, n_act


def agregar_columnas_seguimiento(df: pd.DataFrame, col_p1: str,
                                  col_p2: Optional[str] = None) -> pd.DataFrame:  # CAMBIO
    df = df.copy()

    df["Reportado"] = df[col_p1].apply(
        lambda v: "Sí" if _tiene_dato(v) else "No"
    )

    df["Estado del indicador"] = df[col_p1].apply(
        lambda v: "Reportado" if _tiene_dato(v) else "Pendiente de reporte"
    )
    return df


# ── Escritura con openpyxl ─────────────────────────────────────────────────────

def _estilo_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=C_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=borde, right=borde, bottom=borde)


def escribir_hoja(ws, df: pd.DataFrame, mapa_fechas: dict = None):
    mapa_fechas = mapa_fechas or {}

    for ci, col in enumerate(df.columns, 1):
        val = col
        if col in mapa_fechas and mapa_fechas[col] is not None:
            val = mapa_fechas[col].strftime("%d/%m/%Y")
        cell = ws.cell(row=1, column=ci, value=val)
        _estilo_header(cell)
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(val)) + 2, 10)

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, raw_val in enumerate(row, 1):
            col_name = df.columns[ci - 1]

            # CAMBIO: con dtype=str en la lectura, los Ids ya vienen como string;
            # se normaliza igual por si acaso quedó algún float residual.
            if col_name == "Id":
                val = _id_normalizar(raw_val) if raw_val is not None else None
            elif raw_val is not None and str(raw_val).strip() == "-":
                val = None
            else:
                val = raw_val

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_name == "Reportado":
                if val == "Sí":
                    cell.fill = PatternFill("solid", fgColor=C_SI)
                elif val == "No":
                    cell.fill = PatternFill("solid", fgColor=C_NO)

            elif col_name == "Estado del indicador":
                if val == "Reportado":
                    cell.fill = PatternFill("solid", fgColor=C_SI)
                elif val == "Pendiente de reporte":
                    cell.fill = PatternFill("solid", fgColor=C_PENDIENTE)

            elif col_name == "Revisar":
                # CAMBIO: Revisar es int (0/1); comparar con 1 directamente
                try:
                    if int(val) == 1:
                        cell.fill = PatternFill("solid", fgColor=C_REVISAR1)
                except (TypeError, ValueError):
                    pass

            ancho_actual = ws.column_dimensions[get_column_letter(ci)].width
            contenido_ancho = len(str(val)) + 2 if val is not None else 0
            ws.column_dimensions[get_column_letter(ci)].width = min(
                40, max(ancho_actual, contenido_ancho)
            )

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


# ── Resumen por periodicidad ───────────────────────────────────────────────────

def crear_hoja_resumen(wb, resumen_data: list):
    ws = wb.create_sheet(title="Resumen", index=0)
    headers = ["Periodicidad", "Total indicadores", "Reportados (período actual)",
               "Pendientes de reporte", "% Reporte"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _estilo_header(cell)
        ws.column_dimensions[get_column_letter(ci)].width = max(len(h) + 4, 18)

    for ri, row in enumerate(resumen_data, 2):
        ws.cell(row=ri, column=1, value=row["Periodicidad"])
        ws.cell(row=ri, column=2, value=row["Total"])
        ws.cell(row=ri, column=3, value=row["Reportados"])
        ws.cell(row=ri, column=4, value=row["Pendientes"])
        pct = round(row["Reportados"] / row["Total"] * 100, 1) if row["Total"] else 0
        pct_cell = ws.cell(row=ri, column=5, value=f"{pct}%")
        if pct >= 80:
            pct_cell.fill = PatternFill("solid", fgColor=C_SI)
        elif pct >= 50:
            pct_cell.fill = PatternFill("solid", fgColor=C_PENDIENTE)
        else:
            pct_cell.fill = PatternFill("solid", fgColor=C_NO)


# ── Flujo principal ────────────────────────────────────────────────────────────

def main():
    sep = "=" * 62
    print(sep)
    print("  generar_reporte.py -> Seguimiento_Reporte.xlsx")
    print(sep)

    # ── 1. Leer fuente ────────────────────────────────────────────────────────
    print(f"\n[1] Leyendo fuente: {RUTA_ORIGEN}")
    if not os.path.exists(RUTA_ORIGEN):
        sys.exit(f"    ERROR: No se encontró {RUTA_ORIGEN}")

    # CAMBIO: se llama leer_xlsx (openpyxl) en vez de leer_xls (xlrd)
    df = leer_xlsx(RUTA_ORIGEN)
    print(f"    OK -> {len(df)} filas  |  {len(df.columns)} columnas")

    # ── 2. Columna Revisar ───────────────────────────────────────────────────
    if COLUMNA_REVISAR not in df.columns:
        sys.exit(f"    ERROR: Columna '{COLUMNA_REVISAR}' no encontrada en la fuente.")

    print(f"\n[2] Columna 'Revisar' (basada en '{COLUMNA_REVISAR}')...")
    df = agregar_revisar(df, COLUMNA_REVISAR)
    indicadores_unicos = df["Revisar"].sum()
    print(f"    {indicadores_unicos} indicadores únicos detectados.")

    # ── 3. Identificar columnas de período ───────────────────────────────────
    periodo_cols = [c for c in df.columns if str(c).startswith("Periodo ")]
    if not periodo_cols:
        sys.exit("    ERROR: No se encontraron columnas 'Periodo X' en la fuente.")
    n_periodos = len(periodo_cols)
    print(f"\n[3] Columnas de período: {n_periodos}  ({periodo_cols[0]} … {periodo_cols[-1]})")

    # ── 4. Crear workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 5. Hoja "Seguimiento" ────────────────────────────────────────────────
    print(f"\n[4] Hoja 'Seguimiento' -> {len(df)} filas...")
    ws_seg = wb.create_sheet(title="Seguimiento")
    escribir_hoja(ws_seg, df)
    print("    OK")

    # ── 6. Cargar Kawak (opcional) ───────────────────────────────────────────
    print(f"\n[5] Cargando respaldo Kawak: {RUTA_KAWAK}")
    kawak_lookup = leer_kawak(RUTA_KAWAK)

    # ── 7. Hojas por periodicidad ────────────────────────────────────────────
    # CAMBIO: se filtra None y strings vacíos/nan en la lista de periodicidades
    periodicidades = [
        p for p in df["Periodicidad"].dropna().unique()
        if p and str(p).strip() not in ("", "nan", "NaN", "None")
    ]
    print(f"\n[6] Periodicidades: {periodicidades}")
    resumen_data = []

    for perio in periodicidades:
        df_p = df[df["Periodicidad"] == perio].copy().reset_index(drop=True)
        total = len(df_p)
        print(f"\n    ── {perio} ({total} filas)")

        fechas = get_period_dates(perio, n_periodos)
        mapa = {col: f for col, f in zip(periodo_cols, fechas)}
        if fechas[0]:
            print(f"       Periodo 1 -> {fechas[0].strftime('%d/%m/%Y')}  "
                  f"| Periodo {n_periodos} -> {fechas[-1].strftime('%d/%m/%Y')}")

        col_p1 = periodo_cols[0]
        col_p2 = periodo_cols[1] if n_periodos >= 2 else None
        df_p = agregar_columnas_seguimiento(df_p, col_p1, col_p2)

        fecha_p1 = fechas[0] if fechas else None
        fecha_p2 = fechas[1] if len(fechas) > 1 else None
        df_p, n_kawak = enriquecer_desde_kawak(
            df_p, kawak_lookup, col_p1, col_p2, fecha_p1, fecha_p2
        )

        reportados = (df_p["Reportado"] == "Sí").sum()
        pendientes = (df_p["Estado del indicador"] == "Pendiente de reporte").sum()
        print(f"       Reportados (LMI)            : {reportados - n_kawak}/{total}")
        if kawak_lookup:
            print(f"       Actualizados desde Kawak   : {n_kawak}")
        print(f"       Reportados total           : {reportados}/{total}")
        print(f"       Pendientes de reporte      : {pendientes}/{total}")

        ws = wb.create_sheet(title=perio[:31])
        escribir_hoja(ws, df_p, mapa)

        resumen_data.append({
            "Periodicidad": perio,
            "Total": total,
            "Reportados": int(reportados),
            "Pendientes": int(pendientes),
        })

    # ── 8. Hoja Resumen ──────────────────────────────────────────────────────
    print("\n[7] Hoja 'Resumen'...")
    crear_hoja_resumen(wb, resumen_data)
    print("    OK")

    # ── 9. Guardar ───────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(RUTA_SALIDA), exist_ok=True)
    print(f"\n[8] Guardando en {RUTA_SALIDA}...")
    wb.save(RUTA_SALIDA)
    print(f"    OK -> archivo guardado.")

    print(f"\n{sep}")
    print(f"  Proceso completado exitosamente.")
    print(f"  Archivo: {RUTA_SALIDA}")
    print(sep)


if __name__ == "__main__":
    main()